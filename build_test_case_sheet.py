# -*- coding: utf-8 -*-
"""
Builds a manual QA test case sheet (Excel) for the SMS Ham/Spam API --
covers functional classification across all 4 supported languages and
categories, plus API-level validation/edge cases.
"""
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SMS_Classifier_Manual_Test_Cases.xlsx")
DIVERSE_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "diverse_test_set.csv")

CATEGORY_MAP = {
    "personal": "Personal Conversation",
    "banking_otp": "Banking / OTP (Ham)",
    "legit_url": "Legitimate URL (Ham)",
    "ham_general": "General Ham",
    "promotional": "Promotional Spam",
    "phishing": "Phishing / Malicious URL",
    "mixed_language": "Mixed-Language (Code-Switched)",
}
LANG_MAP = {
    "english": "English", "hinglish": "Hinglish", "hindi": "Hindi",
    "marathi": "Marathi", "mixed": "Mixed (Hindi+Marathi+English)",
}

HEADERS = [
    "TC ID", "Category", "Language", "Test SMS Input", "API Endpoint",
    "Request Body (example)", "Expected Result", "Actual Result",
    "Actual Spam Probability", "Pass / Fail", "Tester Remarks",
    "Tested By", "Date Tested",
]

rows = []  # each: list matching HEADERS (blank for manual-fill columns)
tc_counter = 1


def add_row(category, lang, text, endpoint, expected, body=None):
    global tc_counter
    tc_id = f"TC-{tc_counter:03d}"
    tc_counter += 1
    if body is None:
        body = '{"text": "%s"}' % text.replace('"', '\\"')
    rows.append([tc_id, category, lang, text, endpoint, body, expected,
                 "", "", "", "", "", ""])


# ---- 1. Functional classification cases, pulled from diverse_test_set.csv ----
with open(DIVERSE_CSV, encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    diverse_rows = list(reader)

# Order: group by category then language for a readable sheet
category_order = ["personal", "banking_otp", "legit_url", "ham_general",
                   "promotional", "phishing", "mixed_language"]
lang_order = ["english", "hinglish", "hindi", "marathi", "mixed"]


def sort_key(r):
    cat = r["CATEGORY"]
    lang = r["LANG"]
    return (category_order.index(cat) if cat in category_order else 99,
            lang_order.index(lang) if lang in lang_order else 99)


for r in sorted(diverse_rows, key=sort_key):
    add_row(
        CATEGORY_MAP.get(r["CATEGORY"], r["CATEGORY"]),
        LANG_MAP.get(r["LANG"], r["LANG"]),
        r["TEXT"],
        "POST /predict",
        r["LABEL"],
    )

# ---- 2. API-level validation / edge cases ----
add_row("Edge Case - Validation", "N/A", "", "POST /predict",
        "400 Bad Request - \"text must not be empty\"",
        body='{"text": ""}')
add_row("Edge Case - Validation", "N/A", "   (whitespace only)", "POST /predict",
        "400 Bad Request - \"text must not be empty\"",
        body='{"text": "   "}')
add_row("Edge Case - Validation", "N/A", "(missing \"text\" field)", "POST /predict",
        "422 Unprocessable Entity - validation error", body='{}')
add_row("Edge Case - Validation", "N/A", "(text field is a number, not a string)", "POST /predict",
        "422 Unprocessable Entity - validation error", body='{"text": 12345}')
add_row("Edge Case - Robustness", "English", "Hi", "POST /predict",
        "200 OK - Ham (single short word, observe confidence)")
add_row("Edge Case - Robustness", "N/A", "?", "POST /predict",
        "200 OK - observe behavior (single character, no clear ground truth)")
add_row("Edge Case - Robustness", "N/A", "1234567890 9988776655 00000",
        "POST /predict", "200 OK - observe behavior (digits only)")
add_row("Edge Case - Robustness", "N/A", "@#$%^&*()!!! ..... ????",
        "POST /predict", "200 OK - observe behavior (special characters only)")
add_row("Edge Case - Robustness", "N/A", "🎉🎁💰🔥🤑😍",
        "POST /predict", "200 OK - observe behavior (emoji only)")
add_row("Edge Case - Security", "English",
        "<script>alert('xss')</script> Congratulations! Claim your free prize now: http://fake-prize-claim.xyz/win",
        "POST /predict",
        "200 OK - Spam; must NOT execute/reflect script, just classify text normally")
add_row("Edge Case - Robustness", "English",
        "URGENT ".join(["Your account will be suspended, click here to verify now http://bad-link.xyz/verify. "] * 15),
        "POST /predict",
        "200 OK - Spam (very long/repetitive message; confirms input is truncated safely, not an error)")
add_row("Edge Case - Robustness", "English",
        "Your parcel has shipped, track at https://www.dtdc.in/track/AB123. Also please verify KYC urgently at http://fake-kyc-verify.xyz/click before it expires!",
        "POST /predict",
        "200 OK - Spam (message contains BOTH a legit-looking and a malicious URL; malicious intent should dominate)")
add_row("Edge Case - Language Coverage", "Unsupported (Bengali)",
        "আপনার অ্যাকাউন্ট ব্লক করা হবে, এখনই যাচাই করুন",
        "POST /predict",
        "200 OK - no crash; observe behavior for an unsupported language script")
add_row("Edge Case - Idempotency", "English",
        "Dear Customer, your account will be blocked today. Verify now: http://fake-verify.xyz/click",
        "POST /predict",
        "200 OK - Spam; send this exact request twice, both responses must be identical")
add_row("Edge Case - Endpoint", "N/A", "(no body required)", "GET /health",
        '200 OK - {"status": "ok", "model_config": {...}}', body="(none)")
add_row("Edge Case - Endpoint", "N/A", "(wrong HTTP method on /predict)", "GET /predict",
        "405 Method Not Allowed", body="(none)")
add_row("Edge Case - Batch", "Mixed (Hindi+English)", "3 messages, see body",
        "POST /predict/batch",
        "200 OK - \"results\" list of 3, matching the individual /predict result for each message",
        body=('{"messages": ["Hey, are we still meeting for lunch today?", '
              '"बधाई हो! आपने Rs.50,000 जीता है, अभी दावा करें: http://fake-win.xyz/claim", '
              '"Get flat 60% off on all shoes this weekend only!"]}'))
add_row("Edge Case - Batch", "N/A", "(empty messages list)", "POST /predict/batch",
        "422 Unprocessable Entity - validation error (min 1 message required)",
        body='{"messages": []}')


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# ---- Instructions sheet ----
ws0 = wb.active
ws0.title = "Instructions"
title_font = Font(size=16, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="2F5496")
section_font = Font(size=12, bold=True, color="2F5496")
ws0["A1"] = "SMS Ham/Spam Classifier API - Manual Test Case Sheet"
ws0["A1"].font = title_font
ws0["A1"].fill = title_fill
ws0.merge_cells("A1:D1")
ws0.row_dimensions[1].height = 28

instructions = [
    ("", ""),
    ("How to run the API", ""),
    ("1.", "Install dependencies: pip install -r requirements.txt"),
    ("2.", "Start the server: python main.py  (or: python api.py)"),
    ("3.", "Base URL: http://localhost:8000   |   Interactive docs: http://localhost:8000/docs"),
    ("", ""),
    ("Endpoints", ""),
    ("GET  /health", "Returns {\"status\": \"ok\", ...} if the server and model loaded correctly."),
    ("POST /predict", "Body: {\"text\": \"<sms text>\"}  ->  {\"label\": \"Ham\"|\"Spam\", \"is_spam\": bool, "
                       "\"spam_probability\": float, \"ham_probability\": float}"),
    ("POST /predict/batch", "Body: {\"messages\": [\"<sms 1>\", \"<sms 2>\", ...]}  ->  {\"results\": [...]}"),
    ("", ""),
    ("How to execute each test case", ""),
    ("1.", "Copy the 'Request Body (example)' cell for the row into Postman / curl / the /docs Swagger UI."),
    ("2.", 'Example curl: curl -X POST http://localhost:8000/predict -H "Content-Type: application/json" -d "<body>"'),
    ("3.", "Record the API's actual label and spam_probability in the 'Actual Result' / 'Actual Spam Probability' columns."),
    ("4.", "Mark 'Pass' if Actual Result matches Expected Result, otherwise 'Fail'. Use 'Blocked' if the API errored "
           "unexpectedly, or 'Not Executed' if not yet run."),
    ("5.", "For 'Edge Case - Robustness' rows there is no strict right/wrong label -- just record what the API "
           "returns and confirm it responds with HTTP 200 and no crash."),
    ("", ""),
    ("Sheet contents", ""),
    ("Test Cases", "116 functional classification cases (Personal, Banking/OTP, Legitimate URL, General Ham, "
                    "Promotional Spam, Phishing/Malicious URL, Mixed-Language) across English, Hinglish, Hindi "
                    "and Marathi, plus 18 API-level edge/validation cases."),
    ("Summary", "Auto-calculated pass/fail counters once you fill in the 'Pass / Fail' column on the Test Cases sheet."),
]
r = 3
for left, right in instructions:
    if right == "" and left != "":
        ws0.cell(row=r, column=1, value=left).font = section_font
    else:
        ws0.cell(row=r, column=1, value=left).font = Font(bold=True)
        ws0.cell(row=r, column=2, value=right)
    r += 1
ws0.column_dimensions["A"].width = 22
ws0.column_dimensions["B"].width = 110
for row in ws0.iter_rows(min_row=3, max_row=r - 1, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# ---- Test Cases sheet ----
ws = wb.create_sheet("Test Cases")
header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

category_colors = {
    "Personal Conversation": "DDEBF7",
    "Banking / OTP (Ham)": "E2EFDA",
    "Legitimate URL (Ham)": "FFF2CC",
    "General Ham": "F2F2F2",
    "Promotional Spam": "FCE4D6",
    "Phishing / Malicious URL": "F8CBAD",
    "Mixed-Language (Code-Switched)": "D9D2E9",
}

for i, row_data in enumerate(rows, start=2):
    category = row_data[1]
    fill_color = category_colors.get(category, "FFFFFF")
    if category.startswith("Edge Case"):
        fill_color = "FFE699"
    for col, val in enumerate(row_data, start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=(col in (4, 6, 7, 11)))
        c.fill = PatternFill("solid", fgColor=fill_color)

widths = {1: 10, 2: 26, 3: 16, 4: 55, 5: 16, 6: 45, 7: 34, 8: 16, 9: 20, 10: 14, 11: 30, 12: 14, 13: 14}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Dropdown validation for Actual Result and Pass/Fail columns
dv_actual = DataValidation(type="list", formula1='"Ham,Spam,Error/Exception,Other"', allow_blank=True)
dv_actual.error = "Please choose from the list"
dv_actual.prompt = "Select the label the API actually returned"
ws.add_data_validation(dv_actual)
dv_actual.add(f"H2:H{len(rows) + 1}")

dv_status = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Not Executed"', allow_blank=True)
dv_status.error = "Please choose from the list"
dv_status.prompt = "Select the outcome of this test case"
ws.add_data_validation(dv_status)
dv_status.add(f"J2:J{len(rows) + 1}")

last_row = len(rows) + 1

# ---- Summary sheet ----
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Test Execution Summary"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2.merge_cells("A1:B1")
ws2.row_dimensions[1].height = 28

summary_rows = [
    ("Total Test Cases", f"=COUNTA('Test Cases'!A2:A{last_row})"),
    ("Executed (Pass/Fail/Blocked filled)", f"=COUNTIF('Test Cases'!J2:J{last_row},\"<>\")-COUNTIF('Test Cases'!J2:J{last_row},\"Not Executed\")"),
    ("Passed", f"=COUNTIF('Test Cases'!J2:J{last_row},\"Pass\")"),
    ("Failed", f"=COUNTIF('Test Cases'!J2:J{last_row},\"Fail\")"),
    ("Blocked", f"=COUNTIF('Test Cases'!J2:J{last_row},\"Blocked\")"),
    ("Not Executed", f"=COUNTIF('Test Cases'!J2:J{last_row},\"Not Executed\")+COUNTIF('Test Cases'!J2:J{last_row},\"\")"),
    ("Pass Rate (%)", f"=IFERROR(ROUND(COUNTIF('Test Cases'!J2:J{last_row},\"Pass\")/(COUNTIF('Test Cases'!J2:J{last_row},\"Pass\")+COUNTIF('Test Cases'!J2:J{last_row},\"Fail\"))*100,1),0)"),
]
r = 3
for label, formula in summary_rows:
    ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=formula)
    r += 1
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 20

wb.save(OUT_PATH)

with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "build_test_case_sheet_result.txt"), "w", encoding="utf-8") as f:
    f.write(f"Saved {OUT_PATH} with {len(rows)} test cases.\n")
